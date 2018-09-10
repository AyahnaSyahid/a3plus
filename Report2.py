# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, Color, Fill
from datetime import datetime
from operator import itemgetter
import os
import re

# from typing import List

file_pattern = (
    r'(?P<klien>[^_]+)_'
    r'(?P<nama>[^_]+)_'
    r'(?P<bahan>[^_]+)_'
    r'(?P<str_qty>((?P<page>[\d]+)(?:page|pge|pg)?@)?(?P<kopi>[\d]+)'
    r'(lembar|lbr|lb)?)_?'
    r'(?P<keterangan>.*)\.(?:pdf|tif)$'
)

Matcher = re.compile(file_pattern, re.I)


class Counter():

    def __init__(self):
        self.jumlah_data = 0
        self.base = list()
        self.bahan = dict()

    def summarize_klien(self, _klien):
        pass

    def list_klien(self):
        _all = [x.lower() for x in [d['klien'] for d in self.base]]
        _klien = []
        for a in _all:
            if a not in _klien:
                _klien.append(a)
        return sorted(_klien)

    def list_bahan(self):
        _all = [x.lower() for x in [d['bahan'] for d in self.base]]
        _bahan = []
        for b in _all:
            if b not in _bahan:
                _bahan.append(b)
        return sorted(_bahan)

    def data_filter(self, **kwargs):
        def in_list(a, b):
            for va in a:
                for vb in b:
                    if vb in va:
                        return True
                    if va in vb:
                        return True
            return False

        accepted = []
        for data in self.base:
            acc = None
            for kw in kwargs.keys():
                try:
                    if in_list(kwargs[kw], data[kw].split('_')):
                        acc = data
                    else:
                        acc = None
                        break
                except Exception as e:
                    print(e)
                    break
            if acc is not None:
                accepted.append(acc)
        return accepted

    def add_data(self, _dct):
        if not any(_dct):
            return
        # print("{},{},{}".format(_dct['page'],_dct['kopi'],_dct['keterangan']))
        self.jumlah_data += 1
        self.base.append(_dct)
        try:
            self.bahan[_dct["bahan"]] += int(_dct["kopi"]) * int(_dct["page"])
        except:
            self.bahan[_dct["bahan"]] = int(_dct["kopi"]) * int(_dct["page"])

    def find_data(self, spec):
        ret_val = list()
        for index in self.base:
            for key in index.keys():
                if spec.lower().strip() in key.lower() or spec.lower().strip() \
                        in index[key].lower():
                    print()
                    ret_val.append(index)
                    for k, v in index.items():
                        print("{} = {}".format(k.upper(), v))
        return ret_val

    def remove_data(self, index):
        pass

    def edit_data(self, index):
        pass

    def save_data(self, file_name):
        str_cell = lambda x, y: get_column_letter(x) + str(y)
        wb = Workbook()
        ws = wb.active
        RJUDUL = 1
        RDATE = RJUDUL + 1
        RHEADER = RDATE + 1
        RDATA = RHEADER + 1
        CNO = 1
        CNAMA = CNO + 1
        CNAMAFILE = CNAMA + 1
        CDATA = {}
        _cursor = CNAMAFILE + 1
        for lb in self.list_bahan():
            CDATA[lb] = _cursor
            _cursor += 1
        CQTY = _cursor
        CJUMLAH = CQTY + 1
        CJUDUL = 1  # sampai CJUMLAH
        CDATE = CJUMLAH - 2  # sampai CJUMLAH
        ws[str_cell(CJUDUL, RJUDUL)].value = \
            "LAPORAN HARIAN PRINT A3PLUS"
        for i, COL in zip("NO CUSTOMER NAMAFILE".split(),
                          [CNO, CNAMA, CNAMAFILE]):
            ws[str_cell(COL, RHEADER)].value = \
                i
        for cdkey in CDATA.keys():
            ws[str_cell(CDATA[cdkey], RHEADER)].value = \
                cdkey.upper()[0:6]
        ws[str_cell(CQTY, RHEADER)].value = "QTY"
        ws[str_cell(CJUMLAH, RHEADER)].value = "JUMLAH"
        next_row = RDATA
        self.base = sorted(self.base, key=itemgetter('klien'))
        for data, num in zip(self.base, range(len(self.base))):
            ws[str_cell(CNO, next_row)].value = num + 1
            ws[str_cell(CNAMA, next_row)].value = data['klien']
            ws[str_cell(CNAMAFILE, next_row)].value = \
                data['nama'] + (" ++( " + " ".join(data[
                    'keterangan'].split('_')) + ")" if data[
                    'keterangan'].strip() else '')
            ws[str_cell(CDATA[data['bahan'].lower()], next_row)].value =\
                int(data['page']) * int(data['kopi'])
            ws[str_cell(CQTY, next_row)].value = \
                2 if 'bb' in data['keterangan'].lower().split('_') \
                else 1
            ws[str_cell(CJUMLAH, next_row)].value = \
                "=SUM(" + str_cell(CNAMAFILE + 1, next_row) + ':' + str_cell(
                CQTY - 1, next_row) + ")*" + str_cell(CQTY, next_row)
            next_row += 1
        ws.merge_cells("A1:" + str_cell(CJUMLAH, 1))
        ws.merge_cells(str_cell(CDATE, RDATE) + ":" \
                       + str_cell(CJUMLAH, RDATE)
                       )
        RANGE_HEADER = ws[str_cell(1, RHEADER) + ':' + str_cell(
                CJUMLAH, RHEADER)]
        for row in RANGE_HEADER:
            for cl in row:
                cl.font = Font(bold=True)
                cl.border = Border(left=Side(style='thin'),
                                   right=Side(style='thin'),
                                   top=Side(style='thin'),
                                   bottom=Side(style='thin'))
                cl.fill = PatternFill('solid', Color('00aaaaaa'))
                cl.alignment = Alignment('center','center')
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['B'].width = 20
        TITTLECELL = ws['A1']
        TITTLECELL.font = Font(size=22, bold=True)
        TITTLECELL.alignment = Alignment(horizontal="center", vertical="center")
        DATECELL = ws[str_cell(CDATE, RDATE)]
        DATECELL.alignment = Alignment(horizontal="center", vertical="center")
        DATECELL.value = datetime.now().strftime("%A, %d %B %y")
        csum = ws[str_cell(CJUMLAH, next_row)]
        csum.value = "=SUM(" + str_cell(CJUMLAH,RDATA) +":"+\
                        str_cell(CJUMLAH,next_row -1)+")"
        ws.merge_cells(str_cell(CJUMLAH-3, next_row) + ":" + \
                       str_cell(CJUMLAH-1, next_row))
        ws[str_cell(CJUMLAH-3, next_row)].value = "Jumlah"
        ws[str_cell(CJUMLAH-3, next_row)].fill = PatternFill('solid',Color('00000000'))
        ws[str_cell(CJUMLAH-3, next_row)].font = Font(color="00ffffff", bold=True)
        ws[str_cell(CJUMLAH-3, next_row)].alignment = Alignment(horizontal="right")
        wb.save(file_name)
        
def validate(_file):
    def do_rename(old):
        while True:
            print("Old name = {}".format(old))
            _new = input("Please enter valid name\n")
            try:
                if Matcher.match(_new):
                    os.rename(old, os.path.join(os.path.dirname(old), _new))
                    print("File name changed")
                    break
                else:
                    print("Bad file name : {}".format(_new))
                    continue

            except Exception as ex:
                print("function rename to {} failed".format(_new))
                print(str(ex))

        return os.path.join(os.path.dirname(old), _new)

    def delete(abs_path):
        inp = input("Benarkah ingin menghapus file {}?".format(
            abs_path.split(os.path.sep)[-1]))
        if inp.lower().strip() in "no n".split():
            return True
        else:
            os.remove(abs_path)
        return

    M = Matcher.match(_file['name'])
    if M:
        return _file
    else:
        try:
            _new = do_rename(_file['abs_path'])
            return {'name': os.path.basename(_new), 'abs_path': _new}
        except Exception as e:
            print(str(e))
            return None


def get_file_list(path: str):
    if not os.path.isdir(path):
        raise OSError("Directory doesn't exists")
    ret_val = dict()
    ret_val["abs_path"] = list()
    ret_val["name"] = list()
    for _abs, dir_list, file_list in os.walk(path):
        for fl in [job for job in file_list if job.lower().endswith(".pdf") or
                                               job.lower().endswith(".tif")]:
            ret_val["abs_path"].append(os.path.join(_abs, fl))
            ret_val["name"].append(fl)
    return ret_val


def test():
    # from TTT import filename_gen
    C = Counter()
    base_dir = input("Masukkan nama folder >\n    ")
    flist = get_file_list(base_dir)
    for job in range(len(flist["abs_path"])):
        valid = validate({"name": flist["name"][job], "abs_path": \
            flist["abs_path"][job]})

        C.add_data(Matcher.match(valid['name']).groupdict("1"))
    C.save_data("ok.xlsx")


if __name__ == "__main__":
    test()
#      M = Matcher
#      while True:
#          test_name = input("Masukan nama untuk test\n\t")
#          mtc = M.match(test_name)
#          if mtc:
#              dct = mtc.groupdict("1")
#              for key in [k for k in sorted(dct.keys())]:
#                  print('{} = {}'.format(key.upper(), dct[key]))
#          else:
#              print("Tidak dapat di parse......")
