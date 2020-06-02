# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, Color, Fill
from datetime import datetime
from operator import itemgetter
import os
import re
import pandas as pd

# from typing import List
FILE_ALLOW_EXT = ".pdf"

file_pattern = (
    r'(?P<klien>[^_]+)_'
    r'(?P<nama>[^_]+)_'
    r'(?P<bahan>[^_]+)_'
    r'(?P<str_qty>((?P<page>[\d]+)(?:page|pge|pg)?@)?(?P<kopi>[\d]+)'
    r'(lembar|lbr|lb)?)_?'
    r'(?P<keterangan>.*)\.(?:pdf|tif)$'
)

patern = re.compile(file_pattern, re.I)
isValidName = lambda x: patern.match(x).groupdict("1") or None

if __name__ == "__main__":
    #dapatkan nama direktori
    base_dir = input("Masukan nama direktori : ")
    #list semua nama file
    all_files = []
    for cd, ld, lf in os.walk(base_dir):
        #iterasi nama file dalam base_dir
        for f in [ x for x in lf if x.lower().endswith(FILE_ALLOW_EXT) ] :
            if isValidName(f):
                all_files.append(f)
            else:
                #Lakukan Validasi - Rename
                while not isValidName(f):
                    print("Nama file tidak sesuai patern")
                    old_file = os.path.join(cd, f)
                    new_name = input("Masukan nama file baru :")
                    if (new_name != f) and isValidName(new_name):
                        os.path.rename(old_file, os.path.join(cd, new_name))
                        f = new_name
                        break
                    else:
                        f = new_name
                all_files.append(f)
    #iterasi list file tervalidasi
    data = [ x for x in [ isValidName(f) for f in all_files if isValidName(f) is not None]]
    df = pd.DataFrame(data).astype('int', errors='ignore')
    df['qty'] = df["keterangan"].apply(lambda x: 2 if "bb" in x.lower().split("_") else 1)
    df['jumlah'] = df.kopi.astype(int) * df.page.astype(int) * df.qty.astype(int)
    df.index += 1
    df["No."] = df.index
    df = df[["No.", "klien", "nama", "bahan", "keterangan", "page", "kopi", "qty", "jumlah"]]
    #Type Convert
    df.page = df.page.astype(int)
    df.kopi = df.kopi.astype(int)
    df.columns = [x.upper() for x in ["No.", "Konsumen", "File", "Bahan", "Keterangan", "Halaman", "Banyak", "Side", "Jumlah"]]
    df.to_excel("pppp.xlsx", index=False, startrow=2)
    # Dekorasi
    SHEET_TITLE_RANGE = "A1:I1"
    SHEET_DATE_RANGE = "G2:I2"
    SHEET_COL_RANGE = "A3:I3"
    LAST_ROW = len(df) + 3
    SHEET_DATA_RANGE = "A4:I" + str(LAST_ROW)
    SHEET_DATA_SUM = "I4:I{}".format(LAST_ROW)
    SHEET_SUM_TITLE = "F{}:H{}".format(LAST_ROW + 1, LAST_ROW + 1)
    SHEET_SUM_CELL = "I{}".format(LAST_ROW + 1)
    
    TimesFont = Font(name="Times", size=22, b=True)
    CenterAlignment = Alignment(horizontal='center', vertical='center')
    AllBorder = Border(left=Side(style='thin'),
                                   right=Side(style='thin'),
                                   top=Side(style='thin'),
                                   bottom=Side(style='thin'))
    
    wb = load_workbook("pppp.xlsx")
    ws = wb[wb.sheetnames[0]]
    
    
    #TITLE
    ws.merge_cells(SHEET_TITLE_RANGE)
    ws["A1"] = "LAPORAN A3 PLUS AKSARAJAYA"
    ws["A1"].font = TimesFont
    ws["A1"].alignment = CenterAlignment
    ws.merge_cells(SHEET_DATE_RANGE)
    DATECELL =  ws["G2"]
    DATECELL.alignment = CenterAlignment
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["E"].alignment = CenterAlignment
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10
    
    for r in ws[SHEET_COL_RANGE]:
        for c in r:
            c.fill = PatternFill('solid', Color('00aaaaaa'))
    for r in ws[SHEET_DATA_RANGE]:
        for c in r:
            c.border = Border(left=Side(style='thin'),
                                   right=Side(style='thin'),
                                   top=Side(style='thin'),
                                   bottom=Side(style='thin'))
    cr = 4
    for r in ws[SHEET_DATA_SUM]:
        for c in r:
            ws["I" + str(cr)] = "=F{}*G{}*H{}".format(cr,cr,cr)
        cr += 1
    
    ws.merge_cells("G{}:H{}".format(cr, cr))
    ws["G" + str(cr)].fill = PatternFill('solid', Color("00000000"))
    ws["G" + str(cr)].font = Font(color="00FFFFFF", bold=True, italic=True)
    ws["G" + str(cr)].value = "Jumlah"
    ws["G" + str(cr)].alignment = Alignment(horizontal = "right", vertical = "center")
    ws["I" + str(cr)].border = AllBorder
    ws["I" + str(cr)].value = "=SUM(I4:I{})".format(cr - 1)
    sname = input("Masukan <hari>, tgl bulan tahun  Laporan : ")
    DATECELL.value = sname.strip()
    ws.title = sname.split(",")[1].strip()
    wb.save("ok.xlsx")
    os.remove("pppp.xlsx")
    
    
    
    